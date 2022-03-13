# 2-4
import mod1       #mod1 모듈(PythonLecture클래스 사용을 위함)
import openpyxl   #openpyxl모듈(엑셀파일 부르기 위함)

filename = 'score_table.xlsx'        #엑셀파일에 filename이라는 변수명을 지정
wb_obj=openpyxl.load_workbook(filename)    #엑셀 모듈 여는 함수에 xb_obj라는 변수명 지정
sheet_obj = wb_obj.active                  #시트 객체 지정

for i in range(2, sheet_obj.max_row + 1):  #시트의 데이터를 모두 읽기 위한 반복문
    result = []                            #result는 리스트임을 지정
    for j in range(2, sheet_obj.max_column + 1):
        currentCell_obj = sheet_obj.cell(row=i, column=j)     #각 셀의 데이터를 currentCell_obj으로 지정
        result.append(currentCell_obj.value)                  #각 셀들의 값을 리스트에 추가
    a=mod1.PythonLecture(result[0],result[1],result[2],result[3])   #a는 PythonLecture클래스,리스트 인덱싱을 통해 클래스 변수값 추가
    grade=a.grade()                           #a객체의 grade함수를 grade라는 변수명으로 지정
    Score=a.weightavg()                       #a객체의 weightavg함수를 weightavg라는 변수명으로 지정
    newCell_obj = sheet_obj['F%d'%i]          #F2~F11의 셀 위치지정
    newCell_obj.value = Score                 #각 셀에 최종성적(가중치평균)추가
    newCell_obj = sheet_obj['G%d'%i]          #G2~G11의 셀 위치지정
    newCell_obj.value = grade                 #각 셀에 학점추가

newCell_obj=sheet_obj['F1']                  #F1셀 위치지정
newCell_obj.value='Final Score'              #F1셀에 'Final Score' 문자추가
newCell_obj=sheet_obj['G1']                  #G1셀 위치지정
newCell_obj.value='Final Grade'              #G1셀에 'Final Grade' 문자추가

wb_obj.save('score_table_final.xlsx')        #추가한 데이터를 포함한 데이터를 'score_table_final.xlsx'라는 이름으로 저장
