import classBankAccount    #classBankAccount모듈 부르기
import openpyxl            #openpyxl모듈 부르기
import time                #time 모듈 부르기

filename = 'UserList.xlsx'                                 #엑셀 파일에 filename이라는 변수명 지정
# 사용자 리스트 받고, 등록자 판별
wb_obj = openpyxl.load_workbook(filename)                  #엑셀 모듈 여는 함수에 xb_obj라는 변수명 지정
sheet_obj = wb_obj.active                                  #시트 객체 지정
a = {}                                                     #a 딕셔너리로 지정
for i in range(2, sheet_obj.max_row + 1):                  #행에 대한 반복문
    result = []                                            #result는 리스트임을 지정
    for j in range(1, sheet_obj.max_column + 1):           #열에 대한 반복문
        currentCell_obj = sheet_obj.cell(row=i, column=j)  #각 셀의 데이터를 currentCell_obj으로 지정
        result.append(currentCell_obj.value)               #각 셀들의 값을 리스트에 추가
    a[result[0]] = (result[1], result[2])                  #각 행에 대한 {name:(pw,balane)} 값 딕셔너리에 추가

class Atm(classBankAccount.BankAccount):                   #은행계좌클래스를 상속받은 Atm클래스함수
    def __init__(self):                                    #init함수 오버라이딩
        name = input('Enter the username: ')               #사용자로부터 사용자 이름을 받음
        self.name=name
        if self.name in a.keys():                          #사용자이름이 리스트에 등록되어 있을 때의 조건문
            print('%s 님 환영합니다' %self.name)
            while True:                                                         #반복문
                user_pass = int(input('%s님의 패스워드를 입력하세요 ' % self.name))  #사용자에게 비밀번호를 입력을 받음
                if user_pass == a[self.name][0]:                                #사용자의 비밀번호가 일치했을 때 조건문
                    print('사용자 정보가 확인되었습니다')
                    self.balance = a[self.name][1]                              #balance변수값 지정
                    break                                                       #반복수행 정지
                else:                                                           #비밀번호 일치하지 않았을 경우
                   print('비밀번호가 틀렸습니다.')
     #등록되지 않았을 경우, 사용자 정보 받아 새롭게 등록
        else:
            tell = input('%s 님은 등록되지 않았습니다. 추가하시겠습니까?(yes or no)' % self.name)  # 사용자에게 등록여부받기
            if tell == 'yes':                                                              #등록하겠다고 답했을 때 조건문
                user_pass=int(input('%s 님의 비밀번호를 입력하세요: ' % self.name))             #비밀번호 설정
                self.balance = int(input('%s 님의 초기 잔액을 입력하세요: ' % self.name))       #초기잔액 설정
                a[self.name] = (user_pass, self.balance)                                   #딕셔너리 a에 {name:(pw,balane)} 값 추가
                print('등록이 되었습니다.')
            else:                                                                          # 등록하지 않았을 경우
                print('서비스를 종료합니다.')
    # 입출금 메뉴 반복 수행, 명세표 출력
        if self.name in a.keys():
            while True:
                print('=' * 30)  # 메뉴 출력
                print('원하시는 메뉴를 선택하세요')
                print('1. Deposit \n'
                      '2. Withdraw \n'
                      '3. Check Balance \n'
                      '4. Quit')
                answer = int(input('>>>'))
                if answer == 1:  # 1번인 입금을 선택했을 경우
                    amount = int(input('입금하실 금액을 입력하세요 : '))  # 입금할 금액 사용자로부터 받기
                    self.deposit(amount)  # deposit 함수 불러오기
                    paper = input('명세표를 출력하시겠습니까?(y or n)')  # 명세표 출력여부 사용자로부터 받기
                    if paper == 'y':  # 명세표 출력한다고 했을 때 조건문
                        print('*' * 35)
                        print('             명세표             ')
                        print('거래 시간:', end=' ')  # 거래시간 출력
                        print(time.strftime('%a %b %d %X %Y', time.localtime(time.time())))
                        print('이름: %s' % self.name)  # 거래자명 출력
                        print('입금액: %d' % amount)  # 입급액 출력
                        print('남은 잔액: %d' % self.balance)  # 남은 잔액 출력
                        print('거래해주셔서 감사합니다. -Sookmyung Bank')
                        print('*' * 35)
                elif answer == 2:  # 2번인 출금을 선택했을 경우
                    amount = int(input('출금하실 금액을 입력하세요 : '))  # 출금할 금액 사용자로부터 받기
                    self.withdraw(amount)  # withdraw 함수 불러오기
                    paper = input('명세표를 출력하시겠습니까?(y or n)')  # 명세표 출력여부 사용자로부터 받기
                    if paper == 'y':  # 명세표 출력한다고 했을 때 조건문
                        print('*' * 35)
                        print('             명세표             ')
                        print('거래 시간:', end=' ')  # 거래시간 출력
                        print(time.strftime('%a %b %d %X %Y', time.localtime(time.time())))
                        print('이름: %s' % self.name)  # 거래자명 출력
                        print('출금액: %d' % amount)  # 출금액 출력
                        print('남은 잔액: %d' % self.balance)  # 남은 잔액 출력
                        print('거래해주셔서 감사합니다. -Sookmyung Bank')
                        print('*' * 35)
                elif answer == 3:  # 3번 선택했을 경우
                    print('현재 잔액은 %d입니다.' % self.balance)  # 잔액 알려줌
                else:  # 4번 선택했을 경우
                    print('서비스를 종료합니다.')
                    break  # 서비스 종료(반복문 종료)







